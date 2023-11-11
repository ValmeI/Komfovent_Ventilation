import re
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By

# TODO: need to find out what abbreviations mean, like SP, EP, DX, and so on

# Excel headers
excel_headers = [
    # Columns from url/det.asp
    "Reading Date",
    "Supply temperature - °C",
    "Extract temperature - °C",
    "Outdoor temperature - °C",
    "WT",
    "Panel temperature 1 - °C",
    "Panel temperature 2 - °C",
    "Panel humidity RH 1 - %",
    "Panel humidity RH 2 - %",
    "Supply fan - %",
    "Exhaust fan - %",
    "SP - mV",
    "EP - mV",
    "SFI - %",
    "EFI - %",
    "S1 - mV",
    "S2 - mV",
    "Heat exchanger - %",
    "WC - %",
    "Electric heater - %",
    "DX - %",
    "Air damper - %",
    "Filter clogging - %",
    "Energy saving - %",
    "OH - g/m³",
    "Indoor humidity AH - g/m³",

    # Columns from url/i2.asp
    "Ventilation level i2",
    "Active hours i2",
    "Next Away",

    # Columns from url/i.asp
    "Ventilation level i",
    "Supply temperature - °C",
    "Extract temperature - °C",
    "Outdoor temperature - °C",
    "SP - %",
    "SAF - %",
    "EAF - %",
    "SAFS - %",
    "EAFS - %",
    "Filter clogging - %",
    "Heat exchanger efficiency - %",
    "Heat recovery - W",
    "Power consumption - W",
    "Heating power - W",
    "Specific power Actual",
    "Specific power Day",
    "Consumed energy Day - kWh",
    "Consumed energy Month - kWh",
    "Consumed energy Total - kWh",
    "Heating energy Day - kWh",
    "Heating energy Month - kWh",
    "Heating energy Total - kWh",
    "Recovered energy Day - kWh",
    "Recovered energy Month - kWh",
    "Recovered energy Total - kWh",
    "ST - °C",
    "ET - °C",
    "AQS - %",
    "AQ - %",
    "AHS - %",
    "AH - %",
    "VF"
]

def create_excel(excel_name, sheet_name):
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = sheet_name

    # Freeze 1st row
    sheet1.freeze_panes = 'A2'
    sheet1.append(excel_headers)

    # Add file type
    file_name = excel_name + ".xlsx"
    # Save Excel file
    wb.save(filename=file_name)

def column_width(excel_name):
    file_name = excel_name + ".xlsx"
    workbook_name = file_name
    wb = load_workbook(workbook_name)
    sheet1 = wb.active

    for i, col_value in enumerate(excel_headers, 1):
        # If the column length is very small (less than 5), then give static length of 10, else the length of the column
        if len(col_value) < 5:
            column_extender = 10
        else:
            column_extender = len(col_value)
        # Wants column letter for input, as i. Width input is at the end of it
        sheet1.column_dimensions[get_column_letter(i)].width = column_extender
    wb.save(filename=workbook_name)

def get_vent_stats(komfovent_local_ip, var):
    options = Options()
    options.add_argument("--headless")
    options.add_argument('--no-sandbox')  # Bypass OS security model UPDATE 4.06.2021 problems maybe fixed it
    options.add_argument("--log-level=3")  # Adjust the log level
    options.add_experimental_option('excludeSwitches', ['enable-logging'])  # This line disables the DevTools logging
    #service = Service(executable_path=r"D:\PycharmProjects\chromedriver.exe")
    service = Service(ChromeDriverManager().install())
    service.log_path = "null"  # Disable driver logs
    service.enable_logging = False  # Disable driver logs

    # URL we want to parse: Komfovent local IP
    url = komfovent_local_ip
    driver = webdriver.Chrome(service=service, options=options)
    driver.get(url)

    # Pass file name
    password_file = 'password'

    # Get password from the file, read-only
    open_file = open(password_file + ".txt", 'r', encoding='utf-8')
    for password in open_file:
        # If "CONTROL" is found at the top of the page, then we are still logged in (previous session is still active),
        # so skip the login part and go straight to API. Else, log in.
        if len(driver.find_elements(By.CLASS_NAME, 'name')):
            pass
        else:
            driver.find_element(By.ID, 'i_p').send_keys(password)
            driver.find_element(By.ID, 'b_s').submit()

        # Get data from the direct URL API
        driver.get(url + '//' + var + '.asp')
        content = driver.find_element(By.ID, 'webkit-xml-viewer-source-xml').get_attribute("innerHTML")
        tree = ET.ElementTree(ET.fromstring(content))
        root = tree.getroot()
        md_list = [re.sub(r'[^\d\.]', "", child.text.strip()) 
                   for child in root]
        # Close Chrome
        driver.quit()
        return md_list

def write_to_excel(excel_name, list_of_data):
    # Add file type
    file_name = excel_name + ".xlsx"

    workbook_name = file_name
    wb = load_workbook(workbook_name)
    sheet1 = wb.active
    # New data to write:
    new_data = [list_of_data]

    for info in new_data:
        sheet1.append(info)

    wb.save(filename=workbook_name)

def add_xpos_in_list(var, pos, input_list):
    input_list.insert(pos, var)
    return input_list
